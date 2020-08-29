 # -*- coding: utf-8 -*-

import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from task import Task
from meal import Meal
from heuristic import Heuristic

def createInvervalsForDays(availablePeriods):
    availableIntervals = {}

    for day in availablePeriods.keys():
        availableIntervals[day] = createInvervals(availablePeriods[day][0], availablePeriods[day][1])
    
    return availableIntervals


def createInvervals(startTime, endTime):
    interval = 15
    startTime = startTime.split(':')
    endTime = endTime.split(':')

    intervals = {}
    intervalId = 1
    hour = int(startTime[0])
    minute = int(startTime[1])
    
    while hour < int(endTime[0]):
        iterator = 0
        while minute < 60:
            intervals[(hour, minute)] = 0
            minute += interval
            iterator += 1
            intervalId += 1

        if (minute > 60):
            minute = 15
        else:
            minute = 0

        hour += 1

    iterator = 0
    while minute < int(endTime[1]):
        intervals[(hour, minute)] = 0
        minute += interval
        iterator += 1
        intervalId += 1

    return intervals


def createBusyIntervals(busyPeriods):
    busyIntervals = {}

    busyIntervalId = 1
    for day in busyPeriods.keys():
        tempIntervals = []
        for period in busyPeriods[day].keys():
            tempIntervals.append(createInvervals(busyPeriods[day][period][0], busyPeriods[day][period][1]))

        intervals = []
        busyIntervals[day] = {}
        for interval in tempIntervals:
            for key in interval.keys():
                busyIntervals[day][key] = busyIntervalId
            busyIntervalId += 1

    return busyIntervals


def setBusyIntervals(availableIntervals, busyIntervals):
    for day in busyIntervals.keys():
        for interval in busyIntervals[day].keys():
            availableIntervals[day][interval] = busyIntervals[day][interval]


def createDataStructure(lines):
    numberOfAvailablePeriods = int(lines[0])
    iterator = 1

    availablePeriods = {}
    while iterator <= numberOfAvailablePeriods:
        availablePeriod = lines[iterator].strip('\n').split(' ')
        availablePeriods[availablePeriod[0]] = (availablePeriod[1], availablePeriod[2])
        iterator += 1
    
    availableIntervals = createInvervalsForDays(availablePeriods)

    numberOfBusyPeriods = int(lines[iterator])
    iterator += 1
    busyPeriods = {}
    busyPeriodId = 1
    while iterator <= (numberOfAvailablePeriods + numberOfBusyPeriods + 1):
        busyPeriodsPerDay = lines[iterator].strip('\n').split(' ')
        
        periods = {}
        periodsIterator = 2
        for i in range(int(busyPeriodsPerDay[1])):
            periods[i + 1] = (busyPeriodsPerDay[periodsIterator], busyPeriodsPerDay[periodsIterator + 1])
            periodsIterator += 2
        
        busyPeriods[busyPeriodsPerDay[0]] = periods
        iterator += 1

    busyIntervals = createBusyIntervals(busyPeriods)

    intervals = availableIntervals
    setBusyIntervals(intervals, busyIntervals)

    numberOfTasks = int(lines[iterator])
    iterator += 1
    tasks = {}
    while iterator <= (numberOfAvailablePeriods + numberOfBusyPeriods + numberOfTasks + 2):
        line = lines[iterator].strip('\n').split(' ')
        
        daysItMustBeDone = []
        numberOfDaysItMustBeDone = int(line[2])
        for i in range(numberOfDaysItMustBeDone):
            daysItMustBeDone.append(line[i + 3])

        itMustBeDoneBeforeBusyInterval = False
        busyIntervalThatTheTaskMustBeDoneBefore = 0
        if (line[int(line[2]) + 4] != '-1'):
            itMustBeDoneBeforeBusyInterval = True
            interval = line[int(line[2]) + 6].split(':')
            busyIntervalThatTheTaskMustBeDoneBefore = (line[int(line[2]) + 5], (int(interval[0]), int(interval[1])))

        tasks[line[0]] = Task(line[0], int(line[1]) * 4, daysItMustBeDone, int(line[int(line[2]) + 3]) * 4,
                              itMustBeDoneBeforeBusyInterval, busyIntervalThatTheTaskMustBeDoneBefore)
        iterator += 1

    numberOfMeals = int(lines[iterator])
    iterator += 1
    meals = {}
    while iterator <= (numberOfAvailablePeriods + numberOfBusyPeriods + numberOfTasks + numberOfMeals + 3):
        line = lines[iterator].strip('\n').split(' ')

        meals[line[0]] = Meal(line[1], line[2], line[3])
        iterator += 1
    
    return intervals, busyIntervals, tasks, meals


def printData(intervals, busyIntervals, tasks, meals):
    print('\n\n---------- Busy Intervals ----------\n')
    for day in busyIntervals.keys():
        print(day, ':', busyIntervals[day])

    print('\n\n\n---------- Intervals considering Busy Intervals ----------\n')
    for day in intervals.keys():
        print(day, ':', intervals[day], '\n')

    print('\n\n---------- Tasks ----------\n')
    for task in tasks.keys():
        print(tasks[task].name, ':', tasks[task].workload, tasks[task].daysItMustBeDone, tasks[task].consecutiveMinimumWorkload, tasks[task].itMustBeDoneBeforeBusyInterval, tasks[task].busyIntervalThatTheTaskMustBeDoneBefore)

    print('\n\n\n---------- Meals ----------\n')
    for meal in meals.keys():
        print(meal, ':', meals[meal].possibleIntervals, meals[meal].duration)


def isNumber(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def generateSpreasheet(intervals):
    xlsxFile = Workbook()
    spreadsheet = xlsxFile.active
    spreadsheet.title = "Agenda"

    row = 1
    column = 2
    for day in intervals.keys():
        spreadsheet.cell(row, column, day.upper())
        spreadsheet.cell(row, column).font = Font(bold=True)
        currentCell = spreadsheet.cell(row, column)
        currentCell.alignment = Alignment(horizontal='center')
        row1 = 2
        column1 = 1
        
        for interval in intervals[day].keys():
            hour = interval[0]
            minute = interval[1]
            if (minute == 0):
                minute = str(minute) + '0'
            if (hour < 10):
                hour = '0' + str(interval[0])
            
            spreadsheet.cell(row1, column1, str(hour) + ":" + str(minute))
            spreadsheet.cell(row1, column1).font = Font(bold=True)
            currentCell = spreadsheet.cell(row1, column1)
            currentCell.alignment = Alignment(horizontal='center')

            task = str(intervals[day][interval]).split('_')
            if (len(task) > 1):
                task1 = task
                task = ''
                for i in task1:
                    task += i + ' '
            else:
                task = task[0]

            if isNumber(task):
                if (task == '0'):
                    task = '-'
                else:
                    task = 'Horário ocupado ' + task

            spreadsheet.cell(row1, column, task)
            currentCell = spreadsheet.cell(row1, column)
            currentCell.alignment = Alignment(horizontal='center')
            
            row1 += 1

        column += 1
    
    for col in spreadsheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2) * 1
        spreadsheet.column_dimensions[column].width = adjusted_width

    xlsxFile.save("agenda.xlsx")


def main():
    fileName = ''.join(sys.argv[1:])
    file = open(fileName, "r")
    lines = file.readlines()

    intervals, busyIntervals, tasks, meals = createDataStructure(lines)

    heuristic = Heuristic(intervals, busyIntervals, tasks, meals)
    heuristic.heuristic()

    generateSpreasheet(heuristic.intervals)


if __name__ == "__main__":
    main()